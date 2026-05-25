import base64
import io
import json
import os
import re
import shutil
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl


@dataclass(frozen=True)
class TencentOcrConfig:
    secret_id: str
    secret_key: str
    region: str



def _load_secret(name: str) -> str:
    return (os.getenv(name) or "").strip()



def load_ocr_config_from_env(*, region: str) -> TencentOcrConfig:
    secret_id = _load_secret("TENCENTCLOUD_SECRET_ID") or _load_secret("TENCENT_SECRET_ID")
    secret_key = _load_secret("TENCENTCLOUD_SECRET_KEY") or _load_secret("TENCENT_SECRET_KEY")
    if not secret_id or not secret_key:
        raise RuntimeError("missing OCR credentials in environment")
    return TencentOcrConfig(secret_id=secret_id, secret_key=secret_key, region=region)



def _build_ocr_client(cfg: TencentOcrConfig):
    from tencentcloud.common import credential
    from tencentcloud.common.profile.client_profile import ClientProfile
    from tencentcloud.common.profile.http_profile import HttpProfile
    from tencentcloud.ocr.v20181119 import ocr_client

    cred = credential.Credential(cfg.secret_id, cfg.secret_key)
    http_profile = HttpProfile()
    http_profile.endpoint = "ocr.tencentcloudapi.com"
    client_profile = ClientProfile()
    client_profile.httpProfile = http_profile
    return ocr_client.OcrClient(cred, cfg.region, client_profile)



def general_basic_ocr_image(*, cfg: TencentOcrConfig, image_bytes: bytes, retry: int = 3) -> dict:
    from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
    from tencentcloud.ocr.v20181119 import models

    client = _build_ocr_client(cfg)
    req = models.GeneralBasicOCRRequest()
    req.ImageBase64 = base64.b64encode(image_bytes).decode("ascii")

    last_err: Optional[Exception] = None
    for attempt in range(retry + 1):
        try:
            resp = client.GeneralBasicOCR(req)
            return json.loads(resp.to_json_string())
        except TencentCloudSDKException as e:
            last_err = e
            if attempt >= retry:
                break
            time.sleep(0.8 * (2**attempt))
    raise RuntimeError(f"general basic OCR failed: {last_err}") from last_err



def table_ocr_image(*, cfg: TencentOcrConfig, image_bytes: bytes, mode: str, retry: int = 3) -> dict:
    from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
    from tencentcloud.ocr.v20181119 import models

    client = _build_ocr_client(cfg)
    if mode == "recognize-table":
        req = models.RecognizeTableOCRRequest()
        call = client.RecognizeTableOCR
    elif mode == "recognize-table-accurate":
        req = models.RecognizeTableAccurateOCRRequest()
        call = client.RecognizeTableAccurateOCR
    elif mode == "table":
        req = models.TableOCRRequest()
        call = client.TableOCR
    else:
        raise ValueError(f"unsupported table OCR mode: {mode}")
    req.ImageBase64 = base64.b64encode(image_bytes).decode("ascii")

    last_err: Optional[Exception] = None
    for attempt in range(retry + 1):
        try:
            resp = call(req)
            return json.loads(resp.to_json_string())
        except TencentCloudSDKException as e:
            last_err = e
            if attempt >= retry:
                break
            time.sleep(0.8 * (2**attempt))
    raise RuntimeError(f"table OCR failed: {last_err}") from last_err



def extract_page_text(resp: dict) -> str:
    detections = resp.get("TextDetections") or []
    return "\n".join((x.get("DetectedText") or "").strip() for x in detections if (x.get("DetectedText") or "").strip()).strip()



def should_run_table_ocr(text: str) -> bool:
    if not text:
        return False
    if re.search(r"表\s*\d+\s+\S+", text):
        return True
    return "AQL" in text



def _decode_xlsx_from_ocr_data(resp: dict) -> Optional[bytes]:
    data = resp.get("Data")
    if not isinstance(data, str) or not data.strip():
        return None
    try:
        b = base64.b64decode(data)
    except Exception:
        return None
    return b if b.startswith(b"PK") else None



def _xlsx_to_grid(xlsx_bytes: bytes) -> list[list[str]]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    if not ws.max_row or not ws.max_column:
        return []

    merged_value: dict[tuple[int, int], str] = {}
    for mr in ws.merged_cells.ranges:
        tl = ws.cell(mr.min_row, mr.min_col).value
        tl_text = "" if tl is None else str(tl).strip()
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_value[(r, c)] = tl_text

    grid: list[list[str]] = []
    for r in range(1, ws.max_row + 1):
        row: list[str] = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                v = merged_value.get((r, c), "")
            row.append("" if v is None else str(v).strip())
        grid.append(row)

    while grid and all(not x for x in grid[-1]):
        grid.pop()
    if not grid:
        return []

    rightmost = 0
    for row in grid:
        for i, val in enumerate(row, start=1):
            if val:
                rightmost = max(rightmost, i)
    return [row[:rightmost] for row in grid] if rightmost else []



def _trim_grid_to_table(grid: list[list[str]]) -> list[list[str]]:
    if not grid or len(grid) < 2:
        return grid
    cols = max(len(r) for r in grid)
    grid = [r + [""] * (cols - len(r)) for r in grid]

    def norm(s: str) -> str:
        return re.sub(r"\s+", "", (s or "").strip())

    body_start = None
    for r, row in enumerate(grid):
        first = norm(row[0])
        if re.match(r"^(键|直径|1:100)", first):
            body_start = r
            break
    if body_start is None:
        return grid

    end = len(grid)
    for r in range(body_start, len(grid)):
        first = norm(grid[r][0])
        if re.match(r"^\d+\.\d+", first):
            end = r
            break
    return grid[:end]



def _grid_to_markdown(grid: list[list[str]]) -> str:
    if not grid:
        return ""
    cols = max(len(r) for r in grid)
    if not cols:
        return ""
    rows = [r + [""] * (cols - len(r)) for r in grid]

    def esc(s: str) -> str:
        return (s or "").replace("|", "\\|").replace("\n", " ").strip()

    md = "|" + "|".join(esc(x) for x in rows[0]) + "|\n"
    md += "|" + "|".join(["---"] * cols) + "|\n"
    for row in rows[1:]:
        md += "|" + "|".join(esc(x) for x in row) + "|\n"
    return md.strip()


def table_ocr_to_markdown(resp: dict) -> str:
    xlsx_bytes = _decode_xlsx_from_ocr_data(resp)
    if xlsx_bytes:
        grid = _trim_grid_to_table(_xlsx_to_grid(xlsx_bytes))
        if grid:
            return _grid_to_markdown(grid)
    return ""



def _grid_to_row_chunks(grid: list[list[str]]) -> list[dict]:
    if not grid or len(grid) < 2:
        return []
    cols = max(len(r) for r in grid)
    if cols <= 1:
        return []
    grid = [r + [""] * (cols - len(r)) for r in grid]

    def norm(s: str) -> str:
        s = (s or "").strip()
        if not s:
            return ""
        return re.sub(r"\s+", "", s)

    body_start = None
    for r, row in enumerate(grid):
        first = norm(row[0])
        if re.match(r"^(键|直径|1:100)", first):
            body_start = r
            break
    if body_start is None:
        body_start = 1
    header_rows = max(1, min(body_start, len(grid) - 1))
    blacklist = {"", "GB/T1568-2008", "GBT1568-2008", "GB/T15682008", "表1", "合格质量水平AQL"}

    col_parts: list[list[str]] = [[] for _ in range(cols)]
    for c in range(cols):
        seen: set[str] = set()
        for r in range(header_rows):
            t = norm(grid[r][c])
            if t in blacklist:
                continue
            if t and t not in seen:
                col_parts[c].append(t)
                seen.add(t)
    for c in range(1, cols):
        col_parts[c] = [p for p in col_parts[c] if p not in {"检查项目"}]

    common_prefix = None
    for c in range(1, cols):
        if not col_parts[c]:
            continue
        first = col_parts[c][0]
        if common_prefix is None:
            common_prefix = first
        elif common_prefix != first:
            common_prefix = None
            break
    if common_prefix:
        for c in range(1, cols):
            if col_parts[c] and col_parts[c][0] == common_prefix:
                col_parts[c] = col_parts[c][1:]

    col_labels = []
    for c in range(cols):
        parts = [p for p in col_parts[c] if p]
        label = "-".join(parts).strip("-")
        col_labels.append(label or f"col{c}")

    header = " | ".join(x for x in col_labels if x).strip()
    out: list[dict] = []
    for r in range(header_rows, len(grid)):
        row_name = norm(grid[r][0])
        if not row_name:
            continue
        if re.match(r"^\d+\.\d+", row_name):
            break
        kvs = []
        for c in range(1, cols):
            v = norm(grid[r][c])
            if not v or v in {"-", "—", "一"}:
                continue
            kvs.append(f"{col_labels[c]}={v}")
        row_text = row_name if not kvs else row_name + " | " + " | ".join(kvs)
        out.append({"row_index": r, "header": header, "row_name": row_name, "row_text": row_text})
    return out


def table_ocr_to_row_chunks(resp: dict) -> list[dict]:
    xlsx_bytes = _decode_xlsx_from_ocr_data(resp)
    if xlsx_bytes:
        grid = _trim_grid_to_table(_xlsx_to_grid(xlsx_bytes))
        return _grid_to_row_chunks(grid)
    return []



def process_images_to_artifacts(
    image_paths: list[Path],
    out_dir: Path,
    *,
    cfg: TencentOcrConfig,
    table_mode: str,
    table_policy: str,
) -> dict:
    raw_dir = out_dir / "raw_ocr"
    text_dir = out_dir / "text"
    table_raw_dir = out_dir / "raw_table_ocr"
    table_md_dir = out_dir / "tables_md"
    table_rows_dir = out_dir / "tables_rows"
    for d in (raw_dir, text_dir, table_raw_dir, table_md_dir, table_rows_dir):
        d.mkdir(parents=True, exist_ok=True)

    page_summaries: list[dict] = []
    for image_path in image_paths:
        page_no = int(image_path.stem.split("_")[-1])
        image_bytes = image_path.read_bytes()
        resp = general_basic_ocr_image(cfg=cfg, image_bytes=image_bytes)
        (raw_dir / f"page_{page_no:03d}.json").write_text(
            json.dumps(resp, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        text = extract_page_text(resp)
        (text_dir / f"page_{page_no:03d}.txt").write_text(text + "\n", encoding="utf-8")
        page_out = {"page_no": page_no, "text": text}

        do_table = table_policy == "always" or (table_policy == "auto" and should_run_table_ocr(text))
        if table_mode != "off" and do_table:
            try:
                table_resp = table_ocr_image(cfg=cfg, image_bytes=image_bytes, mode=table_mode)
                (table_raw_dir / f"page_{page_no:03d}.json").write_text(
                    json.dumps(table_resp, ensure_ascii=False, indent=2), encoding="utf-8"
                )
                md = table_ocr_to_markdown(table_resp)
                if md:
                    (table_md_dir / f"page_{page_no:03d}.md").write_text(md + "\n", encoding="utf-8")
                    page_out["table_markdown"] = md
                rows = table_ocr_to_row_chunks(table_resp)
                if rows:
                    with (table_rows_dir / f"page_{page_no:03d}.jsonl").open("w", encoding="utf-8") as f:
                        for item in rows:
                            f.write(json.dumps({"page_no": page_no, **item}, ensure_ascii=False) + "\n")
                    page_out["table_rows"] = rows
            except Exception as e:
                page_out["table_error"] = str(e)
        page_summaries.append(page_out)

    final_path = out_dir / "final_ocr.txt"
    with final_path.open("w", encoding="utf-8") as f:
        for page in sorted(page_summaries, key=lambda x: x["page_no"]):
            f.write(f"===== Page {page['page_no']} =====\n")
            f.write((page.get("text") or "").strip() + "\n\n")
            if page.get("table_markdown"):
                f.write("----- Table Markdown -----\n")
                f.write(page["table_markdown"].strip() + "\n\n")
            if page.get("table_rows"):
                f.write("----- Table Row Chunks -----\n")
                for row in page["table_rows"]:
                    f.write(row["row_text"].strip() + "\n")
                f.write("\n")
    return {"pages": page_summaries, "final_ocr_path": final_path.as_posix()}



def reset_workdir(doc_dir: Path) -> None:
    if doc_dir.exists():
        shutil.rmtree(doc_dir)
    doc_dir.mkdir(parents=True, exist_ok=True)
