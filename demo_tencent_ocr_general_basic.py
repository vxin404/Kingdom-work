import argparse
import base64
import io
import json
import os
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import fitz


@dataclass(frozen=True)
class TencentOcrConfig:
    secret_id: str
    secret_key: str
    region: str


def _iter_page_indices(page_count: int, page_start: int, page_end: int) -> Iterable[int]:
    start = max(1, page_start)
    end = min(page_count, page_end)
    if end < start:
        return []
    return range(start - 1, end)


def render_pdf_pages_to_jpg(
    pdf_path: Path,
    out_dir: Path,
    *,
    page_start: int,
    page_end: int,
    dpi: int,
    jpg_quality: int,
) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    zoom = max(1, int(round(dpi / 72)))
    mat = fitz.Matrix(zoom, zoom)

    doc = fitz.open(pdf_path)
    try:
        image_paths: list[Path] = []
        for page_index in _iter_page_indices(doc.page_count, page_start, page_end):
            page = doc.load_page(page_index)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            jpg_bytes = pix.tobytes("jpg", jpg_quality=jpg_quality)
            out_path = out_dir / f"page_{page_index + 1:03d}.jpg"
            out_path.write_bytes(jpg_bytes)
            image_paths.append(out_path)
        return image_paths
    finally:
        doc.close()


def load_tencent_ocr_config_from_env(*, region: str) -> Optional[TencentOcrConfig]:
    secret_id = os.getenv("TENCENT_SECRET_ID", "").strip()
    secret_key = os.getenv("TENCENT_SECRET_KEY", "").strip()
    if not secret_id or not secret_key:
        return None
    return TencentOcrConfig(secret_id=secret_id, secret_key=secret_key, region=region)


def _build_ocr_client(cfg: TencentOcrConfig):
    from tencentcloud.common import credential
    from tencentcloud.common.profile.client_profile import ClientProfile
    from tencentcloud.common.profile.http_profile import HttpProfile
    from tencentcloud.ocr.v20181119 import ocr_client

    cred = credential.Credential(cfg.secret_id, cfg.secret_key)
    http_profile = HttpProfile()
    http_profile.endpoint = "ocr.tencentcloudapi.com"
    client_profile = ClientProfile()
    client_profile.httpProfile = http_profile
    return ocr_client.OcrClient(cred, cfg.region, client_profile)


def tencent_general_basic_ocr_image(
    *,
    cfg: TencentOcrConfig,
    image_bytes: bytes,
    retry: int = 3,
    sleep_base_s: float = 0.8,
) -> dict:
    from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
    from tencentcloud.ocr.v20181119 import models

    client = _build_ocr_client(cfg)
    req = models.GeneralBasicOCRRequest()
    req.ImageBase64 = base64.b64encode(image_bytes).decode("ascii")

    last_err: Optional[Exception] = None
    for attempt in range(retry + 1):
        try:
            resp = client.GeneralBasicOCR(req)
            return json.loads(resp.to_json_string())
        except TencentCloudSDKException as e:
            last_err = e
            if attempt >= retry:
                break
            time.sleep(sleep_base_s * (2**attempt))
    raise RuntimeError(f"Tencent OCR failed after {retry+1} attempts: {last_err}") from last_err


def tencent_table_ocr_image(
    *,
    cfg: TencentOcrConfig,
    image_bytes: bytes,
    mode: str,
    retry: int = 3,
    sleep_base_s: float = 0.8,
) -> dict:
    from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
    from tencentcloud.ocr.v20181119 import models

    client = _build_ocr_client(cfg)
    req = None
    call = None
    if mode == "recognize-table":
        req = models.RecognizeTableOCRRequest()
        call = client.RecognizeTableOCR
    elif mode == "recognize-table-accurate":
        req = models.RecognizeTableAccurateOCRRequest()
        call = client.RecognizeTableAccurateOCR
    elif mode == "table":
        req = models.TableOCRRequest()
        call = client.TableOCR
    else:
        raise ValueError(f"Unsupported table OCR mode: {mode}")

    req.ImageBase64 = base64.b64encode(image_bytes).decode("ascii")

    last_err: Optional[Exception] = None
    for attempt in range(retry + 1):
        try:
            resp = call(req)
            return json.loads(resp.to_json_string())
        except TencentCloudSDKException as e:
            last_err = e
            if attempt >= retry:
                break
            time.sleep(sleep_base_s * (2**attempt))
    raise RuntimeError(f"Tencent Table OCR failed after {retry+1} attempts: {last_err}") from last_err


def extract_page_text_from_general_basic_ocr(resp: dict) -> str:
    detections = resp.get("TextDetections") or []
    lines: list[str] = []
    for item in detections:
        t = (item.get("DetectedText") or "").strip()
        if t:
            lines.append(t)
    return "\n".join(lines).strip()


def should_run_table_ocr(text: str) -> bool:
    if not text:
        return False
    if re.search(r"表\s*\d+\s+\S+", text):
        return True
    return "AQL" in text


def table_ocr_to_markdown(resp: dict) -> str:
    xlsx_bytes = _decode_xlsx_from_ocr_data(resp)
    if xlsx_bytes:
        grid = _trim_grid_to_table(_xlsx_to_grid(xlsx_bytes))
        if grid:
            return _grid_to_markdown(grid)

    tables = resp.get("TableDetections") or []
    md_parts: list[str] = []

    def esc(s: str) -> str:
        return s.replace("|", "\\|").replace("\n", " ").strip()

    for table in tables:
        raw_cells = table.get("Cells") or []
        cells = [
            c
            for c in raw_cells
            if isinstance(c.get("RowTl"), int)
            and isinstance(c.get("ColTl"), int)
            and isinstance(c.get("RowBr"), int)
            and isinstance(c.get("ColBr"), int)
            and c.get("RowTl") >= 0
            and c.get("ColTl") >= 0
            and c.get("RowBr") >= c.get("RowTl")
            and c.get("ColBr") >= c.get("ColTl")
        ]
        if len(cells) < 8:
            continue

        max_row = 0
        max_col = 0
        for cell in cells:
            rb = cell["RowBr"]
            cb = cell["ColBr"]
            if rb > max_row:
                max_row = rb
            if cb > max_col:
                max_col = cb

        rows = max_row + 1
        cols = max_col + 1
        if rows <= 1 or cols <= 1:
            continue

        grid: list[list[str]] = [["" for _ in range(cols)] for _ in range(rows)]
        for cell in cells:
            text = (cell.get("Text") or "").strip()
            if not text:
                continue
            r = cell["RowTl"]
            c = cell["ColTl"]
            if r < 0 or r >= rows or c < 0 or c >= cols:
                continue
            if grid[r][c]:
                grid[r][c] = (grid[r][c] + " " + text).strip()
            else:
                grid[r][c] = text

        while grid and all(not (x or "").strip() for x in grid[-1]):
            grid.pop()
        if len(grid) <= 1:
            continue

        header = [esc(x) for x in grid[0]]
        body = [[esc(x) for x in row] for row in grid[1:]]

        md = "|" + "|".join(header) + "|\n"
        md += "|" + "|".join(["---"] * cols) + "|\n"
        for row in body:
            md += "|" + "|".join(row) + "|\n"

        titles = table.get("Titles") or []
        title_text = "\n".join(
            (t.get("Text") or "").strip() for t in titles if (t.get("Text") or "").strip()
        ).strip()
        if title_text:
            md_parts.append(title_text + "\n\n" + md.strip())
        else:
            md_parts.append(md.strip())

    return "\n\n".join(md_parts).strip()


def _decode_xlsx_from_ocr_data(resp: dict) -> Optional[bytes]:
    data = resp.get("Data")
    if not isinstance(data, str) or not data.strip():
        return None
    try:
        b = base64.b64decode(data)
    except Exception:
        return None
    return b if b.startswith(b"PK") else None


def _xlsx_to_grid(xlsx_bytes: bytes) -> list[list[str]]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row <= 0 or max_col <= 0:
        return []

    merged_value: dict[tuple[int, int], str] = {}
    for mr in ws.merged_cells.ranges:
        min_row, min_col, max_row_m, max_col_m = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        tl = ws.cell(min_row, min_col).value
        tl_text = "" if tl is None else str(tl).strip()
        for r in range(min_row, max_row_m + 1):
            for c in range(min_col, max_col_m + 1):
                merged_value[(r, c)] = tl_text

    grid: list[list[str]] = []
    for r in range(1, max_row + 1):
        row: list[str] = []
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is None:
                v = merged_value.get((r, c), "")
            s = "" if v is None else str(v).strip()
            row.append(s)
        grid.append(row)

    while grid and all(not x for x in grid[-1]):
        grid.pop()
    if not grid:
        return []

    rightmost = 0
    for row in grid:
        for idx, val in enumerate(row, start=1):
            if val:
                rightmost = max(rightmost, idx)
    if rightmost <= 0:
        return []
    grid = [row[:rightmost] for row in grid]
    return grid


def _grid_to_markdown(grid: list[list[str]]) -> str:
    def esc(s: str) -> str:
        return (s or "").replace("|", "\\|").replace("\n", " ").strip()

    if not grid:
        return ""
    cols = max(len(r) for r in grid)
    if cols <= 0:
        return ""
    norm = [r + [""] * (cols - len(r)) for r in grid]

    header = [esc(x) for x in norm[0]]
    md = "|" + "|".join(header) + "|\n"
    md += "|" + "|".join(["---"] * cols) + "|\n"
    for row in norm[1:]:
        md += "|" + "|".join(esc(x) for x in row) + "|\n"
    return md.strip()


def table_ocr_to_row_chunks(resp: dict) -> list[dict]:
    xlsx_bytes = _decode_xlsx_from_ocr_data(resp)
    if xlsx_bytes:
        grid = _xlsx_to_grid(xlsx_bytes)
        return _grid_to_row_chunks(grid)

    tables = resp.get("TableDetections") or []
    chunks: list[dict] = []
    for table in tables:
        raw_cells = table.get("Cells") or []
        cells = [
            c
            for c in raw_cells
            if isinstance(c.get("RowTl"), int)
            and isinstance(c.get("ColTl"), int)
            and isinstance(c.get("RowBr"), int)
            and isinstance(c.get("ColBr"), int)
            and c.get("RowTl") >= 0
            and c.get("ColTl") >= 0
            and c.get("RowBr") >= c.get("RowTl")
            and c.get("ColBr") >= c.get("ColTl")
        ]
        if len(cells) < 8:
            continue

        max_row = max(c["RowBr"] for c in cells)
        max_col = max(c["ColBr"] for c in cells)
        rows = max_row + 1
        cols = max_col + 1
        if rows <= 1 or cols <= 1:
            continue

        header_row_count = min(3, rows - 1) if rows >= 3 else 1

        def norm(s: str) -> str:
            s = (s or "").strip()
            if not s:
                return ""
            s = re.sub(r"\s+", "", s)
            return s

        def covering_cells(r: int, c: int) -> list[dict]:
            found = []
            for cell in cells:
                if cell["RowTl"] <= r <= cell["RowBr"] and cell["ColTl"] <= c <= cell["ColBr"]:
                    t = (cell.get("Text") or "").strip()
                    if t:
                        found.append(cell)
            return found

        def best_cell_text_for_header(r: int, c: int) -> str:
            cand = covering_cells(r, c)
            if not cand:
                return ""
            exact = [x for x in cand if x["RowTl"] == r]
            chosen = exact if exact else cand
            chosen.sort(key=lambda x: ((x["RowBr"] - x["RowTl"] + 1) * (x["ColBr"] - x["ColTl"] + 1)))
            return norm(chosen[0].get("Text") or "")

        def cell_text_for_body(r: int, c: int) -> str:
            cand = covering_cells(r, c)
            if not cand:
                return ""
            exact = [x for x in cand if x["RowTl"] == r and x["ColTl"] == c]
            chosen = exact if exact else cand
            chosen.sort(key=lambda x: ((x["RowBr"] - x["RowTl"] + 1) * (x["ColBr"] - x["ColTl"] + 1)))
            return norm(chosen[0].get("Text") or "")

        col_parts: list[list[str]] = [[] for _ in range(cols)]
        for c in range(cols):
            seen: set[str] = set()
            for r in range(header_row_count):
                t = best_cell_text_for_header(r, c)
                if t and t not in seen:
                    col_parts[c].append(t)
                    seen.add(t)

        common_prefix = None
        for c in range(1, cols):
            if not col_parts[c]:
                continue
            first = col_parts[c][0]
            if common_prefix is None:
                common_prefix = first
            elif common_prefix != first:
                common_prefix = None
                break
        if common_prefix:
            for c in range(1, cols):
                if col_parts[c] and col_parts[c][0] == common_prefix:
                    col_parts[c] = col_parts[c][1:]

        col_labels = []
        for c in range(cols):
            parts = [p for p in col_parts[c] if p]
            label = "-".join(parts).strip("-")
            col_labels.append(label or f"col{c}")

        header_text = " | ".join(x for x in col_labels if x).strip()

        for r in range(header_row_count, rows):
            row_name = cell_text_for_body(r, 0)
            if not row_name:
                continue
            kvs: list[str] = []
            for c in range(1, cols):
                v = cell_text_for_body(r, c)
                if not v:
                    continue
                key = col_labels[c]
                kvs.append(f"{key}={v}")
            if kvs:
                row_text = row_name + " | " + " | ".join(kvs)
            else:
                row_text = row_name
            chunks.append(
                {
                    "row_index": r,
                    "header": header_text,
                    "row_name": row_name,
                    "row_text": row_text,
                }
            )
    return chunks


def _grid_to_row_chunks(grid: list[list[str]]) -> list[dict]:
    if not grid or len(grid) < 2:
        return []

    rows = len(grid)
    cols = max(len(r) for r in grid)
    if cols <= 1:
        return []
    norm_grid = [r + [""] * (cols - len(r)) for r in grid]

    def norm(s: str) -> str:
        s = (s or "").strip()
        if not s:
            return ""
        return re.sub(r"\s+", "", s)

    body_start = None
    for r in range(rows):
        first = norm(norm_grid[r][0])
        if re.match(r"^(键|直径|1:100)", first):
            body_start = r
            break
    if body_start is None:
        body_start = min(3, rows - 1) if rows >= 4 else 1
    header_row_count = max(1, min(body_start, rows - 1))

    blacklist = {"", "GB/T1568-2008", "GBT1568-2008", "合格质量水平AQL", "表1", "GB/T15682008", "GB/T15682008"}

    col_parts: list[list[str]] = [[] for _ in range(cols)]
    for c in range(cols):
        seen: set[str] = set()
        for r in range(header_row_count):
            t = norm(norm_grid[r][c])
            if t in blacklist:
                continue
            if t and t not in seen:
                col_parts[c].append(t)
                seen.add(t)

    if col_parts and col_parts[0]:
        pass

    for c in range(1, cols):
        col_parts[c] = [p for p in col_parts[c] if p not in {"检查项目"}]

    common_prefix = None
    for c in range(1, cols):
        if not col_parts[c]:
            continue
        first = col_parts[c][0]
        if common_prefix is None:
            common_prefix = first
        elif common_prefix != first:
            common_prefix = None
            break
    if common_prefix:
        for c in range(1, cols):
            if col_parts[c] and col_parts[c][0] == common_prefix:
                col_parts[c] = col_parts[c][1:]

    col_labels: list[str] = []
    for c in range(cols):
        parts = [p for p in col_parts[c] if p]
        label = "-".join(parts).strip("-")
        col_labels.append(label or f"col{c}")

    header_text = " | ".join(x for x in col_labels if x).strip()

    out: list[dict] = []
    for r in range(header_row_count, rows):
        row_name = norm(norm_grid[r][0])
        if not row_name:
            continue
        if re.match(r"^\d+\.\d+", row_name):
            break
        kvs: list[str] = []
        for c in range(1, cols):
            v = norm(norm_grid[r][c])
            if not v:
                continue
            if v in {"-", "—", "一"}:
                continue
            kvs.append(f"{col_labels[c]}={v}")
        row_text = row_name if not kvs else row_name + " | " + " | ".join(kvs)
        out.append(
            {
                "row_index": r,
                "header": header_text,
                "row_name": row_name,
                "row_text": row_text,
            }
        )
    return out


def _trim_grid_to_table(grid: list[list[str]]) -> list[list[str]]:
    if not grid or len(grid) < 2:
        return grid
    cols = max(len(r) for r in grid)
    norm_grid = [r + [""] * (cols - len(r)) for r in grid]

    def norm(s: str) -> str:
        s = (s or "").strip()
        if not s:
            return ""
        return re.sub(r"\s+", "", s)

    body_start = None
    for r in range(len(norm_grid)):
        first = norm(norm_grid[r][0])
        if re.match(r"^(键|直径|1:100)", first):
            body_start = r
            break
    if body_start is None:
        return grid

    end = len(norm_grid)
    for r in range(body_start, len(norm_grid)):
        first = norm(norm_grid[r][0])
        if re.match(r"^\d+\.\d+", first):
            end = r
            break
    return [row[:cols] for row in norm_grid[:end]]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pdf", type=str, required=True)
    parser.add_argument("--region", type=str, default="ap-guangzhou")
    parser.add_argument("--page-start", type=int, default=1)
    parser.add_argument("--page-end", type=int, default=999999)
    parser.add_argument("--dpi", type=int, default=200)
    parser.add_argument("--jpg-quality", type=int, default=85)
    parser.add_argument("--out-dir", type=str, default="artifacts/tencent_ocr_demo")
    parser.add_argument(
        "--table-mode",
        type=str,
        default="off",
        choices=["off", "recognize-table", "recognize-table-accurate", "table"],
    )
    parser.add_argument(
        "--table-policy",
        type=str,
        default="auto",
        choices=["auto", "always", "never"],
    )
    args = parser.parse_args()

    pdf_path = Path(args.pdf).expanduser().resolve()
    if not pdf_path.exists():
        raise FileNotFoundError(pdf_path)

    out_dir = Path(args.out_dir).expanduser().resolve()
    images_dir = out_dir / "images"
    raw_dir = out_dir / "raw_ocr"
    text_dir = out_dir / "text"
    table_raw_dir = out_dir / "raw_table_ocr"
    table_md_dir = out_dir / "tables_md"
    table_rows_dir = out_dir / "tables_rows"
    raw_dir.mkdir(parents=True, exist_ok=True)
    text_dir.mkdir(parents=True, exist_ok=True)
    table_raw_dir.mkdir(parents=True, exist_ok=True)
    table_md_dir.mkdir(parents=True, exist_ok=True)
    table_rows_dir.mkdir(parents=True, exist_ok=True)

    print(f"PDF: {pdf_path}")
    print(f"Output: {out_dir}")
    print("Rendering PDF pages to JPG...")
    image_paths = render_pdf_pages_to_jpg(
        pdf_path,
        images_dir,
        page_start=args.page_start,
        page_end=args.page_end,
        dpi=args.dpi,
        jpg_quality=args.jpg_quality,
    )
    print(f"Rendered {len(image_paths)} pages.")

    cfg = load_tencent_ocr_config_from_env(region=args.region)
    if cfg is None:
        print("Missing credentials. Set env vars and rerun:")
        print("  export TENCENT_SECRET_ID='YOUR_SECRET_ID'")
        print("  export TENCENT_SECRET_KEY='YOUR_SECRET_KEY'")
        print("Optional:")
        print("  export TENCENT_REGION='ap-guangzhou'")
        return 2

    final_pages: list[dict] = []

    for image_path in image_paths:
        page_no = int(image_path.stem.split("_")[-1])
        image_bytes = image_path.read_bytes()
        resp = tencent_general_basic_ocr_image(cfg=cfg, image_bytes=image_bytes)
        raw_path = raw_dir / f"page_{page_no:03d}.json"
        raw_path.write_text(json.dumps(resp, ensure_ascii=False, indent=2), encoding="utf-8")

        text = extract_page_text_from_general_basic_ocr(resp)
        text_path = text_dir / f"page_{page_no:03d}.txt"
        text_path.write_text(text + "\n", encoding="utf-8")

        preview = text[:200].replace("\n", " ")
        print(f"[page {page_no}] chars={len(text)} preview={preview!r}")

        page_out: dict = {"page_no": page_no, "text": text}

        if args.table_mode != "off" and args.table_policy != "never":
            do_table = args.table_policy == "always" or should_run_table_ocr(text)
            if do_table:
                try:
                    table_resp = tencent_table_ocr_image(
                        cfg=cfg, image_bytes=image_bytes, mode=args.table_mode
                    )
                    table_raw_path = table_raw_dir / f"page_{page_no:03d}.json"
                    table_raw_path.write_text(
                        json.dumps(table_resp, ensure_ascii=False, indent=2), encoding="utf-8"
                    )
                    md = table_ocr_to_markdown(table_resp)
                    if md:
                        (table_md_dir / f"page_{page_no:03d}.md").write_text(
                            md + "\n", encoding="utf-8"
                        )
                        md_preview = md.replace("\n", " ")[:200]
                        print(f"[page {page_no}] table_md_preview={md_preview!r}")
                    else:
                        print(f"[page {page_no}] table_ocr_done_no_table_detections")

                    row_chunks = table_ocr_to_row_chunks(table_resp)
                    if row_chunks:
                        rows_path = table_rows_dir / f"page_{page_no:03d}.jsonl"
                        with rows_path.open("w", encoding="utf-8") as f:
                            for item in row_chunks:
                                item_out = {"page_no": page_no, **item}
                                f.write(json.dumps(item_out, ensure_ascii=False) + "\n")
                        page_out["table_rows"] = row_chunks
                    if md:
                        page_out["table_markdown"] = md
                except Exception as e:
                    print(f"[page {page_no}] table_ocr_failed: {e}")

        final_pages.append(page_out)

    final_pages.sort(key=lambda x: x["page_no"])
    final_path = out_dir / "final_ocr.txt"
    with final_path.open("w", encoding="utf-8") as f:
        for page in final_pages:
            page_no = page["page_no"]
            f.write(f"===== Page {page_no} =====\n")
            f.write(page.get("text", "").strip() + "\n\n")
            if page.get("table_markdown"):
                f.write("----- Table Markdown -----\n")
                f.write(page["table_markdown"].strip() + "\n\n")
            if page.get("table_rows"):
                f.write("----- Table Row Chunks -----\n")
                for row in page["table_rows"]:
                    f.write(row["row_text"].strip() + "\n")
                f.write("\n")

    print(f"Wrote final OCR document: {final_path}")
    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
